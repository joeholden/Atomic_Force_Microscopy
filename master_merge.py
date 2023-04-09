import numpy as np
import pandas as pd
import os
import openpyxl
from collections import defaultdict
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import date


# Modify the 2 Lines Below:

# What is the path to the folder containing today's data? End in a forward slash
# Eg. "C:/Users/joema/Desktop/Ghazi_AFM/033023/"
PARENT_DIR = "C:/Users/joema/Desktop/Ghazi_AFM/033023/"

# What is the path to the folder you want the Excel Workbook Saved to? End in a forward slash
EXCEL_PATH = "C:/Users/joema/Desktop/Ghazi_AFM/"

# What is the path to the folder you want the plot images saved to? End in a forward slash
PLOT_PATH = "C:/Users/joema/Desktop/Ghazi_AFM/"


def retrieve_modulus_column(path_to_xls):
    """
    Takes in the .xls file from the AFM instrument
    This file is not actually an .xls file so pd.read_excel will not work
    Instead parse the lines of the raw file and collect the data in the fourth column with a tab delimeter
    Return a pandas Series object containing the Young's Modulus column in kPa and the date of the experiment
    """

    with open(path_to_xls) as f:
        lines = f.readlines()

    young_modulus = []  # kPa
    for i in range(7, len(lines)):  # Data starts in line 7
        line = lines[i]
        line_split_list = line.split('\t')
        if line_split_list != ['\n']:  # Line between data and stats info
            modulus_value = line_split_list[3]  # Young's Modulus is in column 3
        else:
            modulus_value = np.NAN

        try:
            young_modulus.append(float(modulus_value) * 1000)
        except ValueError:
            young_modulus.append(np.NAN)  # This happens because there is an empty row before the stats begin

    young_modulus = pd.Series(young_modulus)
    date = lines[0]  # Report date in line 0
    date = date[date.find('(') + 1:date.find(')')].split(' ')[0]  # Find text between () and split off the timestamp

    return date, young_modulus


wb = openpyxl.Workbook()
ws = wb.active

DATA = defaultdict()

for root, dirs, files in os.walk(PARENT_DIR):
    for file in files:
        if file.endswith(".xls"):
            samples_directory = root.split('/')[-1].split(os.sep)
            sample_number = samples_directory[0].split("_")[0]  # todo use regular expression
            condition = samples_directory[0].split("_")[1]
            try:
                time = samples_directory[0].split("_")[2]
            except IndexError:  # error for the baseline naming
                time = '0'
            spot_number = samples_directory[1].split("_")[0]
            tissue = samples_directory[1].split("_")[1]  # todo use regular expression

            acquisition_date, modulus_series = retrieve_modulus_column(os.path.join(root, file))
            ident = sample_number + "_" + condition + "_" + time

            if sample_number not in DATA.keys():
                DATA[sample_number] = defaultdict(list)
                DATA[sample_number][condition].append([spot_number, tissue, acquisition_date, ident, modulus_series])
            else:
                DATA[sample_number][condition].append([spot_number, tissue, acquisition_date, ident, modulus_series])

for sample in DATA.keys():
    d = list(DATA[sample].items())
    sample_start_col = ws.max_column + 1
    for condition_data in d:  # condition_data: ('cmp', [['spot1', 'PPS'], ['spot2', 'PPS'], ['spot4', 'GL']])
        number_of_spots = len(condition_data[1])

        if ws["A2"].value is not None:
            start_col = 1
        else:
            start_col = ws.max_column + 1

        ws.cell(row=2, column=start_col).value = condition_data[0].upper()
        end_col = get_column_letter(start_col + number_of_spots - 1)
        range_cells = get_column_letter(start_col) + "2:" + end_col + "2"
        ws.merge_cells(range_cells)
        ws.cell(row=2, column=start_col).alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=start_col).font = Font(bold=True)

        for spot in condition_data[1]:
            ws.cell(row=3, column=start_col).value = spot[0].title()
            ws.cell(row=4, column=start_col).value = spot[1].upper()

            modulus_data = spot[4]
            modulus_data = pd.DataFrame(modulus_data, index=None)

            rows = dataframe_to_rows(modulus_data)
            for index, row in enumerate(rows):
                if index != 0 and index != 1:
                    row = list(row)
                    ws.cell(row=row[0] + 5, column=start_col, value=row[1])

            start_col += 1
    sample_end_col = ws.max_column
    ws.cell(row=1, column=sample_start_col).value = sample.title()
    ws.merge_cells(get_column_letter(sample_start_col) + "1:" + get_column_letter(sample_end_col) + "1")
    ws.cell(row=1, column=sample_start_col).alignment = Alignment(horizontal='center')
    ws.cell(row=1, column=sample_start_col).font = Font(bold=True)


ws.cell(row=2, column=1).value = 'Sample Number'
ws.cell(row=2, column=1).font = Font(bold=True)
ws.cell(row=3, column=1).value = 'Spot Number'
ws.cell(row=3, column=1).font = Font(bold=True)
ws.cell(row=4, column=1).value = 'Tissue Type'
ws.cell(row=4, column=1).font = Font(bold=True)
ws.cell(row=262, column=1).value = 'AVG'
ws.cell(row=262, column=1).font = Font(bold=True)
ws.cell(row=263, column=1).value = 'STDEV'
ws.cell(row=263, column=1).font = Font(bold=True)
ws.cell(row=264, column=1).value = 'MAX'
ws.cell(row=264, column=1).font = Font(bold=True)
ws.cell(row=265, column=1).value = 'MIN'
ws.cell(row=265, column=1).font = Font(bold=True)

ws.column_dimensions['A'].width = 15

try:
    ws.title = acquisition_date
except NameError:
    pass

try:
    wb.save(EXCEL_PATH + f"RESULTS_{acquisition_date}.xlsx")
except NameError:
    wb.save(EXCEL_PATH + f"RESULTS_{date.today()}.xlsx")
